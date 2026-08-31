"""Export a `ForceMatrix` in the team's spreadsheet layout.

PORT NOTE (forcesSim, phase 4)
----------------------------------------
Ported from `nfrsim.suspension.export` (NFRFullVehicleSim), with the sheet
set trimmed to match the team's actual reference workbook
(`NFR27ComponentForces.xlsx`) rather than the original's full set:

    KEPT     Key, the seven component sheets, the four "bump + <case>" sheets,
             Bump Swept, Envelope -- 14 sheets, Key FIRST.
    DROPPED  Provenance, Checks -- present in the original but not in the
             team's reference file. `ForceMatrix.provenance` and `.check()`
             still exist and still run (nothing about computing them is
             skipped); only the two sheets that would render them are gone.
             A caller that wants them back can add the two blocks from the
             original `nfrsim.suspension.export.workbook_bytes` verbatim --
             both read straight off `matrix`, no other change needed.

WHAT THIS MIRRORS
-----------------
`Suspension Component Forces NFR26` has one sheet per component, and within
each sheet one block per load case: a header naming the case, then a small
table of attachment point against Fx / Fy / Fz / Net. That is the shape a
structures reviewer already knows how to read, so the export reproduces it
rather than emitting a tidy long-format table that would be easier to write and
harder to use.

UNITS AND FRAME
---------------
lbf by default, matching the source spreadsheet.

The frame defaults to ISO 8855 (x forward, y left, z up), which is the same as
the sheet's "x = forward, y = axial, z = up". `frame="solidworks"` rewrites
every component in the CAD triad (x left, y up, z forward) so a load can be
pasted into a study without anyone permuting it by hand -- which is the step
that puts a cornering load into the braking direction and is invisible once
it has happened.

BOTH the unit and the frame are written into every sheet header rather than
assumed. A force table with no unit on it is a trap; one with no frame on it is
the same trap, and the two exports are otherwise byte-identical in shape, so
nothing but the header distinguishes them once the file is on someone's desk.
"""

from __future__ import annotations

import csv
import io
from typing import Dict, List, Optional

from ..core.frames import get_report_frame, to_report_frame
from ..core.units import LBF_TO_N
from .force_matrix import ForceMatrix
from .glossary import glossary_rows

__all__ = ["SHEET_NAMES", "write_workbook", "write_csv", "workbook_bytes"]

#: Component key -> spreadsheet sheet name, in the order the sheets appear.
SHEET_NAMES: Dict[str, str] = {
    "upright": "Uprights",
    "wishbone_upper": "Wishbones Upper",
    "wishbone_lower": "Wishbones Lower",
    "rod": "Push-Pullrods",
    "tie_rod": "Tie Rods",
    "rocker": "Rockers",
    "hub": "Hubs",
}

_COLUMNS = ("Point", "Fx", "Fy", "Fz", "Net")


def _rows_for(matrix: ForceMatrix, axle: str, component: str, case: str,
              lbf: bool, frame) -> List[dict]:
    return [r for r in matrix.rows(axle, component, lbf=lbf, frame=frame)
            if r["Case"] == case]


def _result_rows(result, component: str, lbf: bool, frame) -> List[dict]:
    """`_table`-shaped rows for one component of one solved case."""
    return [{"Point": point, **v}
            for point, v in result.forces.table(component, lbf=lbf,
                                                frame=frame).items()]


#: Excel's hard limits on a sheet name, worth honouring explicitly rather than
#: discovering via a corrupt workbook.
_SHEET_NAME_MAX = 31
_SHEET_NAME_BAD = set(r"[]:*?/\\")


def _combined_sheet_title(case_name: str) -> str:
    """"bump + outside_cornering" -> "Bump + Outside Cornering".

    Derived from the case name rather than a lookup table, so renaming or
    adding a `LoadCase` cannot leave a tab labelled with the old one.
    """
    title = case_name.replace("_", " ").title()
    title = "".join(c for c in title if c not in _SHEET_NAME_BAD)
    return title[:_SHEET_NAME_MAX]


def write_workbook(matrix: ForceMatrix, path, lbf: bool = True,
                   frame=None) -> None:
    """Write the full workbook to `path` (.xlsx).

    `frame` is the reporting triad: `None`/`"iso8855"` for the internal frame,
    `"solidworks"` for x left, y up, z forward.
    """
    with open(path, "wb") as fh:
        fh.write(workbook_bytes(matrix, lbf=lbf, frame=frame))


def workbook_bytes(matrix: ForceMatrix, lbf: bool = True, frame=None) -> bytes:
    """The workbook as bytes, for a download button or a file write."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:                    # pragma: no cover - env dependent
        raise ImportError(
            "xlsx export needs openpyxl (`pip install openpyxl`). Use "
            "`write_csv` for a dependency-free export.") from e

    unit = "lbf" if lbf else "N"
    frame = get_report_frame(frame)
    wb = Workbook()
    wb.remove(wb.active)

    head_font = Font(bold=True, size=12)
    case_font = Font(bold=True, size=11, color="FFFFFF")
    case_fill = PatternFill("solid", fgColor="44546A")
    col_font = Font(bold=True)
    col_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _table(ws, row: int, title: str, rows: List[dict],
               extra: Optional[str] = None) -> int:
        """Write one titled block; return the next free row."""
        ws.cell(row=row, column=1, value=title).font = case_font
        ws.cell(row=row, column=1).fill = case_fill
        for c in range(2, len(_COLUMNS) + 2):
            ws.cell(row=row, column=c).fill = case_fill
        row += 1
        if extra:
            ws.cell(row=row, column=1, value=extra).font = Font(italic=True, size=9)
            row += 1

        for i, name in enumerate(_COLUMNS, start=1):
            c = ws.cell(row=row, column=i, value=name)
            c.font = col_font
            c.fill = col_fill
            c.border = border
            c.alignment = Alignment(horizontal="center")
        row += 1

        for r in rows:
            ws.cell(row=row, column=1, value=r["Point"]).border = border
            for i, key in enumerate(("Fx", "Fy", "Fz", "Net"), start=2):
                c = ws.cell(row=row, column=i, value=round(float(r[key]), 1))
                c.number_format = "0.0"
                c.border = border
            row += 1
        return row + 1

    # -- key -----------------------------------------------------------
    # FIRST, not last: it is the sheet a reader who did not build this
    # workbook needs before any of the numbers, and it is where the team's
    # reference file puts it.
    ws = wb.create_sheet("Key")
    ws.cell(row=1, column=1, value="Key").font = head_font
    row = 3
    last_section = None
    for section, term, meaning in glossary_rows(frame):
        if section != last_section:
            if last_section is not None:
                row += 1
            c = ws.cell(row=row, column=1, value=section)
            c.font, c.fill = case_font, case_fill
            for i in (2, 3):
                ws.cell(row=row, column=i).fill = case_fill
            row += 1
            for i, name in enumerate(("Term", "Meaning"), start=1):
                c = ws.cell(row=row, column=i, value=name)
                c.font, c.fill, c.border = col_font, col_fill, border
            row += 1
            last_section = section
        ws.cell(row=row, column=1, value=term).border = border
        c = ws.cell(row=row, column=2, value=meaning)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A4"

    # -- one sheet per component, one block per headline case --------------
    for component, sheet_name in SHEET_NAMES.items():
        if component not in matrix.components:
            continue
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1,
                value=f"{sheet_name} — forces in {unit}, "
                      f"{frame.header}").font = head_font
        subtitle = ("Each vector is the force applied TO this component at "
                    "the named point.")
        if component == "rod":
            # Name the ACTUAL actuation per axle. "Rods" alone leaves a reader
            # guessing which of the two they are looking at, and this car runs a
            # different one at each end. Taken from the geometry, not assumed.
            kinds = []
            for axle in ("front", "rear"):
                res = matrix.get(axle, matrix.case_names[0]) if matrix.case_names else None
                if res is not None and res.rod_kind:
                    kinds.append(f"{axle} = {res.rod_kind}")
            if kinds:
                subtitle += ("  |  This is the PUSH/PULLROD, not the tie rod: "
                             + ", ".join(kinds) + ".")
        ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=9)
        row = 4
        for axle in ("front", "rear"):
            for case in matrix.case_names:
                rows = _rows_for(matrix, axle, component, case, lbf, frame)
                if not rows:
                    continue
                result = matrix.get(axle, case)
                row = _table(ws, row, f"{axle.upper()} — {case}", rows,
                             extra=result.note if result and result.note else None)

        ws.column_dimensions["A"].width = 12
        for i in range(2, len(_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12
        ws.freeze_panes = "A4"

    # -- bump-on-top cases, one tab each ----------------------------------
    #
    # These get their own tabs rather than extra blocks on the component sheets
    # so that each combination can be read as a WHOLE CASE -- every component at
    # once -- and so the component sheets stay at the five headline cases. The
    # Envelope shows only whichever combination wins a given point; these tabs
    # show all of them.
    for case_name in matrix.combined_case_names:
        ws = wb.create_sheet(_combined_sheet_title(case_name))
        ws.cell(row=1, column=1,
                value=f"{case_name} — forces in {unit}, "
                      f"{frame.header}").font = head_font
        note = next((matrix.get_combined(a, case_name).note
                     for a in ("front", "rear")
                     if matrix.get_combined(a, case_name) is not None), "")
        ws.cell(row=2, column=1,
                value=(f"{note}. The increment is the bump's peak contact load "
                       f"MINUS the static load, so weight is not counted twice."
                       if note else "")).font = Font(italic=True, size=9)
        row = 4
        for axle in ("front", "rear"):
            result = matrix.get_combined(axle, case_name)
            if result is None:
                continue
            for component, sheet_name in SHEET_NAMES.items():
                if component not in result.forces.components:
                    continue
                row = _table(ws, row, f"{axle.upper()} — {sheet_name}",
                             _result_rows(result, component, lbf, frame))

        ws.column_dimensions["A"].width = 12
        for i in range(2, len(_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12
        ws.freeze_panes = "A4"

    # -- the swept bump, so the Envelope stays traceable -------------------
    #
    # EVERY NUMBER ON THE ENVELOPE SHEET MUST BE FINDABLE ON ANOTHER SHEET.
    #
    # One sheet, every swept winner, with the case string that the Envelope
    # cites. Different points win at different instants and often at different
    # speeds, which is the whole reason the sweep exists, so this is a list of
    # per-point results rather than a case block like the other tabs.
    if matrix.swept:
        ws = wb.create_sheet("Bump Swept")
        ws.cell(row=1, column=1,
                value=f"Swept bump — worst instant per attachment point, "
                      f"{unit}, {frame.header}").font = head_font
        ws.cell(row=2, column=1,
                value="The bump walked over every instant of every speed in "
                      "load_cases.bump.speed_sweep_m_s, alone and superposed on "
                      "each steady case. Each point takes its own worst instant "
                      "-- they do not coincide, which is why no single bump "
                      "table can stand in for this. These are the rows the "
                      "Envelope cites as 'swept'."
                ).font = Font(italic=True, size=9)
        row = 4
        for axle in ("front", "rear"):
            for component in matrix.components:
                pts = matrix.swept.get(axle, {}).get(component, {})
                if not pts:
                    continue
                ws.cell(row=row, column=1,
                        value=f"{axle.upper()} — "
                              f"{SHEET_NAMES.get(component, component)}"
                        ).font = case_font
                ws.cell(row=row, column=1).fill = case_fill
                for c in range(2, 7):
                    ws.cell(row=row, column=c).fill = case_fill
                row += 1
                for i, name in enumerate(("Point", "Case", "Fx", "Fy", "Fz", "Net"),
                                         start=1):
                    c = ws.cell(row=row, column=i, value=name)
                    c.font, c.fill, c.border = col_font, col_fill, border
                row += 1
                for point, r in pts.items():
                    # `matrix.swept` is stored in NEWTONS in the internal ISO
                    # frame -- it never goes through `forces.table`, which is
                    # where BOTH the lbf conversion and the reporting-frame
                    # permutation normally happen. Do both here or the sheet
                    # silently reports newtons under an lbf header, and ISO
                    # components under a SolidWorks one.
                    scale = (1.0 / LBF_TO_N) if lbf else 1.0
                    vec = to_report_frame(
                        [r["Fx"], r["Fy"], r["Fz"]], frame) * scale
                    ws.cell(row=row, column=1, value=point).border = border
                    ws.cell(row=row, column=2, value=r["Case"]).border = border
                    values = (*(float(x) for x in vec), float(r["Net"]) * scale)
                    for i, value in enumerate(values, start=3):
                        c = ws.cell(row=row, column=i, value=round(value, 1))
                        c.border, c.alignment = border, Alignment(horizontal="right")
                    row += 1
                row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 34
        for i in range(3, 7):
            ws.column_dimensions[get_column_letter(i)].width = 12
        ws.freeze_panes = "A4"

    # -- envelope ---------------------------------------------------------
    ws = wb.create_sheet("Envelope")
    ws.cell(row=1, column=1,
            value=f"Worst case per attachment point — {unit}, "
                  f"{frame.header}").font = head_font
    ws.cell(row=2, column=1,
            value="Maximum per point, never a sum of cases. Steady cases are "
                  "alternatives; 'bump + <case>' rows are the bump's vertical "
                  "increment superposed on that case, which is the one "
                  "combination that physically co-occurs. See the Key sheet."
            ).font = Font(italic=True, size=9)
    row = 4
    for axle in ("front", "rear"):
        for component in matrix.components:
            env = matrix.envelope(axle, component, lbf=lbf, frame=frame)
            if not env:
                continue
            ws.cell(row=row, column=1,
                    value=f"{axle.upper()} — {SHEET_NAMES.get(component, component)}"
                    ).font = case_font
            ws.cell(row=row, column=1).fill = case_fill
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = case_fill
            row += 1
            headers = ("Point", "Case", "Fx", "Fy", "Fz", "Net")
            for i, name in enumerate(headers, start=1):
                c = ws.cell(row=row, column=i, value=name)
                c.font, c.fill, c.border = col_font, col_fill, border
            row += 1
            for r in env:
                ws.cell(row=row, column=1, value=r["Point"]).border = border
                ws.cell(row=row, column=2, value=r["Case"]).border = border
                for i, key in enumerate(("Fx", "Fy", "Fz", "Net"), start=3):
                    c = ws.cell(row=row, column=i, value=round(float(r[key]), 1))
                    c.number_format = "0.0"
                    c.border = border
                row += 1
            row += 1
    ws.column_dimensions["A"].width = 12
    # Wide enough for the combined case names ("bump + outside_cornering").
    ws.column_dimensions["B"].width = 26
    for i in range(3, 7):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_csv(matrix: ForceMatrix, path=None, lbf: bool = True,
              frame=None) -> str:
    """Long-format CSV of everything. Dependency-free fallback.

    `frame` is the reporting triad -- see `workbook_bytes`. It goes in the
    header row next to the unit, for the same reason: this file has no sheet
    title to carry it, and a bare Fx/Fy/Fz column is not self-describing.

    Returns the CSV text; also writes it to `path` when given.
    """
    frame = get_report_frame(frame)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Axle", "Case", "Component", "Point", "Fx", "Fy", "Fz", "Net",
                f"unit={'lbf' if lbf else 'N'}", f"frame={frame.key}",
                frame.legend])
    # `all_results`, NOT `results` -- the latter holds only the five headline
    # cases, so the CSV used to omit every bump-on-top combination without
    # saying so. A long-format export that quietly drops rows is worse than one
    # that never had them, because nothing about the file looks incomplete.
    for r in matrix.all_results():
        for component in r.forces.components:
            for point, v in r.forces.table(component, lbf=lbf,
                                           frame=frame).items():
                w.writerow([r.axle, r.case_name, component, point,
                            f"{v['Fx']:.3f}", f"{v['Fy']:.3f}",
                            f"{v['Fz']:.3f}", f"{v['Net']:.3f}", "", "", ""])
    text = buf.getvalue()
    if path is not None:
        with open(path, "w", newline="") as fh:
            fh.write(text)
    return text
