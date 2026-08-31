"""One Excel workbook in, one `cfg` dict out -- the team-facing replacement for YAML.

WHY THIS EXISTS
----------------
Every physics function in this package (`build_vehicle_params`,
`build_corner_geometry`, `extract_components`, `corner_loads_for_case`, ...)
takes a plain nested `cfg` dict shaped like a parsed vehicle YAML -- see
`core/config.py`'s module docstring. This module is the ONLY new adapter in
the whole port: it reads ONE spreadsheet (front geometry + rear geometry +
every vehicle parameter the forces pipeline needs) and builds that exact same
`cfg` shape, so every physics function downstream runs completely unmodified.
Nothing in `core/config.py`, `suspension/`, `tire/` or `vehicle/` was touched
to make this work.

PORT NOTE (forcesSim, phase 5) -- replaces `vehicle_csv.py`
-------------------------------------------------------------
The original team-facing loader (`vehicle_csv.py`, still in git history) read
ONE flat CSV with a `Section` column telling a geometry row from a param row
apart -- workable, but a poor fit for a human editing it: every row shares
nine columns regardless of kind, so most of any given row is blank, and
Excel's own row/column semantics (freeze panes, per-sheet width, a dropdown
for `Axle`) are unavailable in a single flat table. This module reads the same
information from a two-tab .xlsx instead, with everything downstream of
`_read_workbook` -- the parameter registry, the per-axle expansion, the
wheel-centre derivation, the temp hardpoint-sheet writer `build_corner_geometry`
reads -- carried over unchanged from `vehicle_csv.py`. CSV input is no longer
accepted; there is exactly one input format now, not two to keep in sync.

THE WORKBOOK FORMAT
---------------------
Two sheets:

  * **Geometry** -- `Axle, Name, X_mm, Y_mm, Z_mm, Notes`. One row per
    suspension hardpoint. `Axle` is `front` or `rear`, `Name` is the point
    name (`IB_UppLead`, `OB_UppPnt`, ...), `X_mm`/`Y_mm`/`Z_mm` are its
    coordinates in the SAME SolidWorks convention as the original hardpoint
    sheets this project has always used (millimetres, right-side geometry,
    origin at the wheelbase midpoint/centreline/ground plane; +Y up, +Z
    forward, -X to the right). These rows can be copied straight out of a CAD
    export.

  * **Parameters** -- `Axle, Name, Value, Unit`. One row per scalar vehicle
    parameter. `Name` is a dotted key (see `_GLOBAL_PARAMS`/`_PER_AXLE_PARAMS`
    below for the full list), `Value` is the number or word, `Unit` documents
    it, `Axle` is blank for a whole-car parameter or `front`/`rear` for one
    that differs by axle.

Column order does not matter -- both sheets are read by HEADER NAME, from row
1, case-insensitively. A row with a blank `Name` on either sheet is ignored
(use it for a spacer). The left/right mirroring that produces FL/RL from these
front/rear sheets, and the wheel-centre point (which is fully determined by
wheelbase, track and loaded radius -- see `_wheel_centre_m` -- so it is not a
workbook input at all), live in `run_sim.py`, exactly where the old Streamlit
GUI's `app/model.py` put the equivalent logic.
"""

from __future__ import annotations

import atexit
import csv
import shutil
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

__all__ = ["VehicleXlsxError", "load_vehicle_xlsx"]

AXLES = ("front", "rear")


class VehicleXlsxError(ValueError):
    """The input workbook is missing a sheet, a row, or a row has an
    unusable value.

    Distinct from a plain exception so the message can always point at the
    exact sheet/`Axle`/`Name` a team member needs to add or fix, the same way
    `core.config.ConfigError` points at a dotted YAML path.
    """


# ---------------------------------------------------------------------------
# The parameter registry -- every `Parameters` row this loader understands.
# ---------------------------------------------------------------------------
# (dotted cfg path, value type, unit shown in error messages)
#
# This list is deliberately short: it is exactly what `build_vehicle_params`,
# `build_corner_geometry`, `build_brake_bias`, `build_drive_split`,
# `build_caliper_on_upright`, `extract_components` and the bump case
# (`build_quarter_car`/`build_bump_envelope`) read for the forces pipeline
# (see core/config.py and suspension/component_forces.py) -- nothing from the
# old YAML that only fed the dropped GUI tab (gear ratios, motor curves, tire-
# fit coefficients) is here.
_GLOBAL_PARAMS: Tuple[Tuple[str, Any, str], ...] = (
    # `Any` (not `type`) for the middle element -- every row is a `float` or a
    # `str` except `"float_list"`, a string tag `_coerce` special-cases rather
    # than a type object.
    ("mass.vehicle_kg", float, "kg"),
    ("mass.driver_kg", float, "kg"),
    ("mass.cg_height_m", float, "m"),
    ("mass.weight_dist_front", float, "fraction, 0-1"),
    ("dimensions.wheelbase_m", float, "m"),
    ("dimensions.track_front_m", float, "m"),
    ("dimensions.track_rear_m", float, "m"),
    ("chassis.torsional_stiffness_Nm_per_deg", float, "N*m/deg"),
    ("tires.vertical_rate_Npm", float, "N/m"),
    ("aero.ClA", float, "lift coefficient x area, m^2"),
    ("aero.CdA", float, "drag coefficient x area, m^2"),
    ("aero.balance_front", float, "fraction of downforce on the front axle"),
    ("aero.cp_height_m", float, "m"),
    ("aero.air_density_kg_m3", float, "kg/m^3"),
    ("brakes.bias_front", float, "fraction of brake force at the front axle"),
    ("brakes.caliper_mount", str, "upright | inboard"),
    ("powertrain.architecture", str,
     "awd_quad_hub | rwd_hub | rwd_diff | fwd_diff | awd_hybrid"),
    ("bearings.inner_offset_m", float,
     "m, fallback only -- the sheet's IB_Bearing point wins if present"),
    ("bearings.outer_offset_m", float,
     "m, fallback only -- the sheet's OB_Bearing point wins if present"),
    ("bearings.axially_located", str, "inner | outer | split | UNCONFIRMED"),
    # Bump structural case (core.config.build_quarter_car / build_bump_envelope,
    # vehicle/quarter_car.py) -- needed only for export_workbook.py's bump
    # case, not for the four steady-state cases run_sim.py writes.
    ("dampers.rate_reference", str,
     "at_damper -- the only reference point build_quarter_car knows how to "
     "convert"),
    ("dampers.knee_velocity_m_per_s", float,
     "m/s at the damper, where the digressive curve's slope changes"),
    ("load_cases.bump.height_m", float, "m, half-sine bump height"),
    ("load_cases.bump.length_m", float, "m, half-sine bump longitudinal extent"),
    ("load_cases.bump.speed_sweep_m_s", "float_list",
     "m/s, semicolon-separated speeds the bump is swept over"),
)

_PER_AXLE_PARAMS: Tuple[Tuple[str, Any, str], ...] = (
    ("suspension.topology", str, "double_wishbone | five_link"),
    ("suspension.actuation", str, "pushrod | pullrod"),
    ("suspension.rod_pickup", str,
     "upper_wishbone | lower_wishbone | upright"),
    ("suspension.static_camber_deg", float, "deg, negative = top-in"),
    ("suspension.static_toe_deg", float, "deg, sheet's own sign convention"),
    ("suspension.spring_rate_Npm", float, "N/m, at the spring"),
    ("suspension.arb_wheelrate_Npm", float, "N/m at the wheel, 0 if no ARB"),
    ("mass.unsprung_kg_per_corner", float, "kg, one corner"),
    ("wheels.loaded_radius_m", float, "m"),
    ("brakes.rotor_effective_radius_m", float, "m"),
    ("dampers.high_speed.bump_Ns_per_m", float, "N/(m/s), at the damper"),
    ("dampers.high_speed.rebound_Ns_per_m", float, "N/(m/s), at the damper"),
    ("dampers.low_speed.bump_Ns_per_m", float, "N/(m/s), at the damper"),
    ("dampers.low_speed.rebound_Ns_per_m", float, "N/(m/s), at the damper"),
)

_GEOMETRY_HEADER = ["SolidWorks", "X", "Y", "Z", "Note"]
# "notes"/"unit" are documentation, never read by the physics -- see
# `_header_index`. Only the columns a value is actually looked up by are
# required.
_GEOMETRY_REQUIRED_COLUMNS = ("axle", "name", "x_mm", "y_mm", "z_mm")
_PARAMETERS_REQUIRED_COLUMNS = ("axle", "name", "value")


def _set(cfg: Dict[str, Any], dotted: str, value: Any) -> None:
    node = cfg
    parts = dotted.split(".")
    for part in parts[:-1]:
        node = node.setdefault(part, {})
    node[parts[-1]] = value


def _coerce(raw: str, kind, *, sheet: str, axle: str, name: str, unit: str):
    raw = raw.strip()
    if kind is float:
        try:
            return float(raw)
        except ValueError:
            raise VehicleXlsxError(
                f"{sheet} row Axle={axle!r} Name={name!r}: "
                f"Value {raw!r} is not a number ({unit}).") from None
    if kind == "float_list":
        # Semicolon-separated, not comma-separated -- a bare list of numbers
        # in one cell reads cleanly either way in Excel, but semicolons match
        # what a copy/paste out of this cell into another tool expects, and
        # avoid any ambiguity with a locale that uses comma decimals.
        try:
            return [float(x) for x in raw.split(";") if x.strip()]
        except ValueError:
            raise VehicleXlsxError(
                f"{sheet} row Axle={axle!r} Name={name!r}: "
                f"Value {raw!r} is not a semicolon-separated list of numbers "
                f"({unit}).") from None
    return raw


def _header_index(row, *, required: Tuple[str, ...], sheet: str, path: Path
                  ) -> Dict[str, int]:
    """Column name (lowercased) -> 0-based index, from a sheet's first row.

    By NAME rather than fixed position, so reordering columns in the
    spreadsheet does not silently misread the sheet -- and so a missing
    REQUIRED column fails loudly, naming the sheet, rather than reading the
    wrong one under an old header. Only `required` is enforced: a purely
    decorative column (`Notes`, `Unit`) that got trimmed while someone was
    tidying the sheet is not a reason to refuse the whole file -- `_cell`
    already reads back "" for any column absent from the returned index.
    """
    got = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        got[str(cell).strip().lower()] = i
    missing = [c for c in required if c not in got]
    if missing:
        raise VehicleXlsxError(
            f"{path}: sheet {sheet!r} is missing required column(s) {missing} "
            f"in its header row (row 1). Have: {sorted(got)}.")
    return got


def _cell(row, idx: Dict[str, int], name: str) -> str:
    if name not in idx:
        return ""
    i = idx[name]
    if i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v).strip()


def _read_workbook(path: Path) -> Tuple[Dict[str, List[dict]], Dict[Tuple[str, str], str]]:
    """One pass over the two sheets -> (geometry rows per axle, param values
    by (axle, name), axle='' for a global param)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)

    for sheet in ("Geometry", "Parameters"):
        if sheet not in wb.sheetnames:
            raise VehicleXlsxError(
                f"{path}: missing a {sheet!r} sheet (have {wb.sheetnames}).")

    geometry: Dict[str, List[dict]] = {"front": [], "rear": []}
    ws = wb["Geometry"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise VehicleXlsxError(f"{path}: 'Geometry' sheet has no header row.")
    idx = _header_index(header, required=_GEOMETRY_REQUIRED_COLUMNS,
                        sheet="Geometry", path=path)
    for row in rows:
        axle = _cell(row, idx, "axle").lower()
        name = _cell(row, idx, "name")
        if not name and not axle:
            continue                                    # blank spacer row
        if axle not in AXLES:
            raise VehicleXlsxError(
                f"{path}: 'Geometry' row {name!r} has Axle={axle!r}; must be "
                f"'front' or 'rear'.")
        if not name:
            continue
        geometry[axle].append({
            "name": name,
            "x": _cell(row, idx, "x_mm"),
            "y": _cell(row, idx, "y_mm"),
            "z": _cell(row, idx, "z_mm"),
            "note": _cell(row, idx, "notes"),
        })

    params: Dict[Tuple[str, str], str] = {}
    ws = wb["Parameters"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise VehicleXlsxError(f"{path}: 'Parameters' sheet has no header row.")
    idx = _header_index(header, required=_PARAMETERS_REQUIRED_COLUMNS,
                        sheet="Parameters", path=path)
    for row in rows:
        axle = _cell(row, idx, "axle").lower()
        name = _cell(row, idx, "name")
        if not name:
            continue                                    # blank spacer row
        if axle and axle not in AXLES:
            raise VehicleXlsxError(
                f"{path}: 'Parameters' row {name!r} has Axle={axle!r}; must "
                f"be blank, 'front' or 'rear'.")
        params[(axle, name)] = _cell(row, idx, "value")

    wb.close()
    return geometry, params


def _write_axle_geometry_csv(rows: List[dict], path: Path, *, axle: str) -> None:
    """Re-serialize this axle's geometry rows in the exact format
    `suspension.geometry.load_hardpoints_csv` already reads -- so that
    trusted, unmodified parser is what actually builds the corner, not a
    second implementation of the same parsing living in this file."""
    if not rows:
        raise VehicleXlsxError(
            f"no geometry rows found for Axle={axle!r} -- add rows to the "
            f"'Geometry' sheet with Axle={axle}, Name=<point>, "
            f"X_mm/Y_mm/Z_mm=...")
    with path.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(_GEOMETRY_HEADER)
        for r in rows:
            w.writerow([r["name"], r["x"], r["y"], r["z"], r["note"]])


def _wheel_centre_m(wheelbase_m: float, track_m: float, loaded_radius_m: float,
                    *, front: bool) -> List[float]:
    """The right-hand wheel centre, ISO metres -- derived, not entered.

    Exactly the arithmetic the original YAML configs' `wheels.centre_m` was
    hand-filled with (x = +/-wheelbase/2, y = -track/2, z = loaded radius; see
    `configs/nfr27_baseline.yaml`'s own comment for the derivation and the CAD
    cross-check). Deriving it here removes an entire category of hand-entry
    that could silently disagree with the wheelbase/track this same file
    already states.
    """
    x = wheelbase_m / 2.0 if front else -wheelbase_m / 2.0
    return [x, -track_m / 2.0, loaded_radius_m]


def load_vehicle_xlsx(path) -> Dict[str, Any]:
    """Parse one unified vehicle workbook into the `cfg` dict shape
    `core.config`'s builders expect.

    Writes the front/rear geometry sheets back out as two temporary
    SolidWorks-format CSVs (cleaned up at process exit) and points
    `suspension.<axle>.hardpoints_file` at them, so `build_corner_geometry`
    and everything built on it run completely unmodified.
    """
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    geometry, raw_params = _read_workbook(path)

    def get(name: str, kind, unit: str, *, axle: str = "") -> Any:
        key = (axle, name)
        if key not in raw_params:
            where = f"Axle={axle}, " if axle else ""
            kind_name = kind if isinstance(kind, str) else kind.__name__
            raise VehicleXlsxError(
                f"missing required row: sheet=Parameters, {where}Name={name} "
                f"(a {kind_name}, {unit}).")
        return _coerce(raw_params[key], kind, sheet="Parameters", axle=axle,
                       name=name, unit=unit)

    cfg: Dict[str, Any] = {"hardpoint_overrides": {}}
    for name, kind, unit in _GLOBAL_PARAMS:
        _set(cfg, name, get(name, kind, unit))

    # Per-axle params each need their OWN cfg path shape (some nest the axle
    # last, e.g. `mass.unsprung_kg_per_corner.front`; some nest it in the
    # middle, e.g. `suspension.front.topology`), so they are set explicitly
    # rather than generated from `_PER_AXLE_PARAMS`' flat dotted names --
    # that tuple exists to document the full input surface in one place, not
    # to be mechanically expanded.
    for axle in AXLES:
        _set(cfg, f"suspension.{axle}.topology",
             get("suspension.topology", str, "double_wishbone | five_link", axle=axle))
        _set(cfg, f"suspension.{axle}.actuation",
             get("suspension.actuation", str, "pushrod | pullrod", axle=axle))
        _set(cfg, f"suspension.{axle}.rod_pickup",
             get("suspension.rod_pickup", str,
                 "upper_wishbone | lower_wishbone | upright", axle=axle))
        _set(cfg, f"suspension.{axle}.static_camber_deg",
             get("suspension.static_camber_deg", float, "deg", axle=axle))
        _set(cfg, f"suspension.{axle}.static_toe_deg",
             get("suspension.static_toe_deg", float, "deg", axle=axle))
        _set(cfg, f"suspension.{axle}.spring_rate_Npm",
             get("suspension.spring_rate_Npm", float, "N/m", axle=axle))
        _set(cfg, f"suspension.{axle}.arb_wheelrate_Npm",
             get("suspension.arb_wheelrate_Npm", float, "N/m", axle=axle))
        _set(cfg, f"mass.unsprung_kg_per_corner.{axle}",
             get("mass.unsprung_kg_per_corner", float, "kg", axle=axle))
        _set(cfg, f"wheels.loaded_radius_m.{axle}",
             get("wheels.loaded_radius_m", float, "m", axle=axle))
        _set(cfg, f"brakes.rotor_effective_radius_m.{axle}",
             get("brakes.rotor_effective_radius_m", float, "m", axle=axle))
        # Damper rates, at the damper (see `dampers.rate_reference` above) --
        # only needed for the bump case (`core.config.build_quarter_car`).
        _set(cfg, f"dampers.high_speed.{axle}.bump_Ns_per_m",
             get("dampers.high_speed.bump_Ns_per_m", float, "N/(m/s), at the damper",
                 axle=axle))
        _set(cfg, f"dampers.high_speed.{axle}.rebound_Ns_per_m",
             get("dampers.high_speed.rebound_Ns_per_m", float, "N/(m/s), at the damper",
                 axle=axle))
        _set(cfg, f"dampers.low_speed.{axle}.bump_Ns_per_m",
             get("dampers.low_speed.bump_Ns_per_m", float, "N/(m/s), at the damper",
                 axle=axle))
        _set(cfg, f"dampers.low_speed.{axle}.rebound_Ns_per_m",
             get("dampers.low_speed.rebound_Ns_per_m", float, "N/(m/s), at the damper",
                 axle=axle))

    # Wheel centre: DERIVED from wheelbase/track/loaded radius, not a workbook
    # field -- see `_wheel_centre_m`.
    wb = cfg["dimensions"]["wheelbase_m"]
    _set(cfg, "wheels.centre_m.front",
         _wheel_centre_m(wb, cfg["dimensions"]["track_front_m"],
                         cfg["wheels"]["loaded_radius_m"]["front"], front=True))
    _set(cfg, "wheels.centre_m.rear",
         _wheel_centre_m(wb, cfg["dimensions"]["track_rear_m"],
                         cfg["wheels"]["loaded_radius_m"]["rear"], front=False))

    # Geometry: write each axle's sheet to a temp file and point
    # `hardpoints_file` at it (absolute path, so it resolves regardless of
    # this workbook's own location).
    tmp_dir = Path(tempfile.mkdtemp(prefix="forcessim_vehicle_xlsx_"))
    atexit.register(shutil.rmtree, tmp_dir, True)
    for axle in AXLES:
        sheet_path = tmp_dir / f"{axle}.csv"
        _write_axle_geometry_csv(geometry[axle], sheet_path, axle=axle)
        _set(cfg, f"suspension.{axle}.hardpoints_file", str(sheet_path))

    return cfg
